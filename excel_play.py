from openpyxl import load_workbook
import pandas as pd


month = 'october'
days_in_month = 31

def replace_values_by_color(file_path, save_path):
    # Load the workbook and select the active sheet
    wb = load_workbook(file_path)
    ws = wb.active

    # Define the color-to-text mapping
    color_mapping = {
        'FFFF0000': 'вдп',  # Red
        'FF0070C0': 'вдр',  # Blue
      #  'FFFFC000': 'оплач',  # Orange checked
        'FF7030A0': 'шпт'  # Violet
    }

    # Iterate over cells in the worksheet
    x = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.fill.fgColor not in x:
                x.append(cell.fill.fgColor)
            if cell.fill :#and isinstance(cell.fill.fgColor, PatternFill):
                cell_color = cell.fill.fgColor.rgb
                # Replace value based on color mapping
                if cell_color in color_mapping:
                    cell.value = color_mapping[cell_color]

    # Save the modified file
    wb.save(save_path)
    return x


# Example usage
colors = replace_values_by_color('oct_colored.xlsx', 'october_replaced.xlsx')

def process_excel_file(file_path):
    # Load the Excel file into a DataFrame, ensuring all columns are read as strings
    df = pd.read_excel(file_path, dtype=str)
    df.columns = df.columns.astype(str)
    print(df.columns)
    # Iterate over rows and apply the logic for columns 2 to 31
    for col in range(1, days_in_month+1):  # From column '2' to '31'
        prev_col = str(col - 1)  # Previous column name as string
        curr_col = str(col)      # Current column name as string

        # Apply the rule: if value in current column is '0' and previous column is '1', set current to '2'
        df.loc[(df[prev_col] != '0') & (df[curr_col] == '0') & (
            df[prev_col].isin(['БЧ','МВГ', 'ОВО', 'Р', 'МЗ', 'ЛЗ','НО'])), curr_col] = '2'

    # Calculate how many '1' and '2' values are in each row and store in a new column "payed_days"
    df['payed_days'] = df.loc[:, '1':str(days_in_month)].apply(
        lambda row: row.isin(['БЧ','МВГ', 'ОВО', 'Р', 'МЗ', 'ЛЗ','НО', '2']).sum(), axis=1)

    # Save the modified DataFrame to a new Excel file
    df.to_excel(f"final_{month}_calc.xlsx", index=False)

# Example usage
process_excel_file(f'{month}_replaced.xlsx')
