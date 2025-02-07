import os
import re
import csv
import ast
from docx import Document
from openpyxl import load_workbook
import pandas as pd
from datetime import datetime



def extract_cyrillic_patterns(doc_path, output_txt_path, output_csv_path):
    # Open the Word document
    doc = Document(doc_path)

    # Define the regex pattern to include all Ukrainian Cyrillic letters
    # - Optional "Capital letter.-" before the main pattern
    # - Apostrophe (’) allowed after the second big letter or between small letters
    # - Whitespace allowed after the capital letter and dot
    pattern = re.compile(r'(?:[А-ЩЬЮЯҐЄІЇ]\.-)?([А-ЩЬЮЯҐЄІЇ])\.\s?([А-ЩЬЮЯҐЄІЇа-щьюяґєії]{1,}(?:’?[а-щьюяґєії]+)?)')

    # Regex to capture content inside tags (e.g., {МВГ} ... {/МВГ})
    tag_pattern = re.compile(r'\{([А-ЯҐЄІЇ]+)\}([\s\S]+?)\{\/\1\}', re.DOTALL)

    # Concatenate all paragraph text into a single string to allow for cross-paragraph tag searches
    full_text = '\n'.join([paragraph.text for paragraph in doc.paragraphs])

    # List to store the extracted patterns along with the tag
    extracted_patterns = []

    # Search for text inside tags across the entire document text
    tag_matches = tag_pattern.findall(full_text)

    for tag, content in tag_matches:
        # Search for the Cyrillic pattern inside the content of each tag
        matches = pattern.findall(content)
        for match in matches:
            # Reconstruct the full match and pair it with the tag
            full_match = f"{match[0]}.{match[1]}"
            extracted_patterns.append((full_match, tag))

    # Write the extracted patterns to the output text file
    with open(output_txt_path, 'w', encoding='utf-8') as txt_file:
        for pattern, tag in extracted_patterns:
            txt_file.write(f"{pattern} ({tag})\n")

    # Write the extracted patterns to the output CSV file
    with open(output_csv_path, 'w', newline='', encoding='utf-8') as csv_file:
        writer = csv.writer(csv_file)
        for pattern, tag in extracted_patterns:
            writer.writerow([pattern, tag])


def process_folder(input_folder, output_folder, month):
    # Ensure the output folder exists
    os.makedirs(output_folder, exist_ok=True)

    # Loop through each .docx file in the input folder
    for filename in os.listdir(input_folder):
        if filename.endswith('.docx'):
            doc_path = os.path.join(input_folder, filename)

            # Extract the day number from the filename
            day_number = int(filename.split('_')[1].split('.')[0])  # Extract day from format "БН_04.08.2024.docx"

            # Define output paths for .txt and .csv files with the format "aug_{N}"
            output_txt_path = os.path.join(output_folder, f"{month}_{day_number}.txt")
            output_csv_path = os.path.join(output_folder, f"{month}_{day_number}.csv")

            # Apply the extraction function
            extract_cyrillic_patterns(doc_path, output_txt_path, output_csv_path)
            print(f"Processed and saved: {filename} as {month}_{day_number}")


def analyze_dataframes(extracted, roster):
    # 1) Lists of items present in both but not in 1:1 relation
    # Group by 'name' and 'name_initials' and count occurrences
    extracted_grouped = extracted['name'].value_counts()
    roster_grouped = roster['name_initials'].value_counts()

    # Find items that are present in both dataframes
    common_items = set(extracted_grouped.index).intersection(set(roster_grouped.index))

    # Filter for items where there's more than one occurrence in at least one of the dataframes
    not_one_to_one = [item for item in common_items if roster_grouped[item] > 1]

    # 2) List of items that are 1:1 (no duplicates in either dataframe)
    one_to_one = [item for item in common_items if extracted_grouped[item] >= 1 and roster_grouped[item] == 1]

    # 3) List of items present in 'roster' but absent in 'extracted'
    missing_in_extracted = list(set(roster['name_initials']) - set(extracted['name']))

    # 4) List of items present in 'extracted' but absent in 'roster'
    missing_in_roster = list(set(extracted['name']) - set(roster['name_initials']))

    return not_one_to_one, one_to_one, missing_in_extracted, missing_in_roster


def update_roster_with_one_to_one(roster, outputs_folder, days_in_month, month):
    for i in range(0, days_in_month + 1):  # Loop from 1 to the number of days in the month
        # Construct file path
        file_path = os.path.join(outputs_folder, f"{month}_{i}_edited.csv")

        # Check if the file exists
        if os.path.exists(file_path):
            # Read the CSV file into a DataFrame
            extracted = pd.read_csv(file_path, names=['name', 'type'])

            # Apply analyze_dataframes function
            _, one_to_one, _, _ = analyze_dataframes(extracted, roster)

            # Create a new column in the 'roster' DataFrame named i (as a string)
            column_name = str(i)
            roster[column_name] = 0

            # Update the column with the 'type' value where the item is present in one_to_one list
            for name in one_to_one:
                type_value = extracted.loc[extracted['name'] == name, 'type'].values[0]
                roster.loc[roster['name_initials'] == name, column_name] = type_value


def replace_values_by_color(file_path, save_path):
    # Load the workbook and select the active sheet
    wb = load_workbook(file_path)
    ws = wb.active

    # Define the color-to-text mapping
    color_mapping = {
        'FFFF0000': 'вдп',  # Red
        'FF0070C0': 'вдр',  # Blue
      #  'FFFFC000': 'оплач',  # Orange checked
        'FF7030A0': 'шпт'  # Violet
    }

    # Iterate over cells in the worksheet
    x = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.fill.fgColor not in x:
                x.append(cell.fill.fgColor)
            if cell.fill :#and isinstance(cell.fill.fgColor, PatternFill):
                cell_color = cell.fill.fgColor.rgb
                # Replace value based on color mapping
                if cell_color in color_mapping:
                    cell.value = color_mapping[cell_color]

    # Save the modified file
    wb.save(save_path)
    return x


def process_excel_file(file_path, duties, days_in_month):
    # Load the Excel file into a DataFrame, ensuring all columns are read as strings
    df = pd.read_excel(file_path, dtype=str)
    df.columns = df.columns.astype(str)
    print(df.columns)
    # Iterate over rows and apply the logic for columns 2 to 31
    for col in range(1, days_in_month+1):  # From column '2' to '31'
        prev_col = str(col - 1)  # Previous column name as string
        curr_col = str(col)      # Current column name as string

        # Apply the rule: if value in current column is '0' and previous column is '1', set current to '2'
        df.loc[(df[prev_col] != '0') & (df[curr_col] == '0') & (
            df[prev_col].isin(duties)), curr_col] = '2'

    # Calculate how many '1' and '2' values are in each row and store in a new column "payed_days"
    df['payed_days'] = df.loc[:, '1':str(days_in_month)].apply(
        lambda row: row.isin(duties + ['АРБА', '2']).sum(), axis=1)

    # Save the modified DataFrame to a new Excel file
    df.to_excel("with_calculated_days.xlsx", index=False)


def extract_dates(df, month, year):
    # Helper function to convert day, month, and year into 'dd.mm.yyyy' format
    def convert_to_date(day, month, year):
        date_obj = datetime(year, month, day)
        return date_obj.strftime('%d.%m.%Y')

    # Function to extract dates from intervals
    def get_dates(intervals, month, year):
        from_dates = []
        to_dates = []

        for interval in intervals:
            if interval:  # Check if interval is not empty
                from_day = interval[0]  # First element of the interval
                to_day = interval[1]  # Second element of the interval

                # Convert days to date strings
                from_dates.append(convert_to_date(from_day, month, year))
                to_dates.append(convert_to_date(to_day, month, year))

        # Join dates with breaklines ('\n')
        return '\n'.join(from_dates), '\n'.join(to_dates)

    # Apply the get_dates function to each row
    df['from'] = df['intervals'].apply(lambda x: get_dates(x, month, year)[0])
    df['to'] = df['intervals'].apply(lambda x: get_dates(x, month, year)[1])


def squash_intervals(numbers):
    """
    Takes a list of numbers and returns a list of intervals.
    Each interval is represented by the first and last number in the consecutive sequence.
    If a number is not part of a sequence, it will still appear as a pair [number, number].

    Args:
        numbers (list): A list of sorted numbers.

    Returns:
        list: A list of lists, where each list is a pair representing an interval.
    """
    if not numbers:
        return []

    # Initialize the list to hold squashed intervals
    squashed = []

    # Initialize the start of the first interval
    start = numbers[0]

    for i in range(1, len(numbers)):
        # If the current number is not consecutive
        if numbers[i] != numbers[i - 1] + 1:
            # Append the interval to the result as a pair [start, end]
            squashed.append([start, numbers[i - 1]])
            # Update the start of the next interval
            start = numbers[i]

    # Add the final interval as a pair [start, end]
    squashed.append([start, numbers[-1]])

    return squashed


def safe_literal_eval(val):
    try:
        # Parse string with ast.literal_eval
        parsed_val = ast.literal_eval(val)
        # Ensure '{}' is converted to an empty dictionary
        if isinstance(parsed_val, dict):
            return parsed_val
        return parsed_val
    except (ValueError, SyntaxError):
        # For non-parsable strings, return as-is
        return {} if val == '{}' else val


